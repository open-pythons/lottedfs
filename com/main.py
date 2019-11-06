# !/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import threading
from bs4 import BeautifulSoup
import random
import sys
import os
import re
import sqlite3
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import yaml
import atexit
import time
import win32api
import signal


def signal_handler(signal, frame):
    pass


signal.signal(signal.SIGINT, signal_handler)
search_url = 'http://chn.lottedfs.cn/kr/search?comSearchWord={0}&comCollection=GOODS&comTcatCD=&comMcatCD=&comScatCD=&comPriceMin=&comPriceMax=&comErpPrdGenVal_YN=&comHsaleIcon_YN=&comSaleIcon_YN=&comCpnIcon_YN=&comSvmnIcon_YN=&comGiftIcon_YN=&comMblSpprcIcon_YN=&comSort=RANK%2FDESC&comListCount=20&txtSearchClickCheck=Y'
targeturl = 'http://icanhazip.com/'  # 验证ip有效性的指定url
sys.path.append(os.path.abspath(os.path.dirname(__file__) + '/' + '..'))
sys.path.append("..")
yamlPath = 'config.yaml'
_yaml = open(yamlPath, 'r', encoding='utf-8')
cont = _yaml.read()
yaml_data = yaml.load(cont, Loader=yaml.FullLoader)
pattern = re.compile('[0-9]+')


# 返回一个随机的请求头 headers
def getheaders():
    user_agent_list = [
        'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.1 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2226.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.4; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2225.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2225.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2224.3 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/40.0.2214.93 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2062.124 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2049.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 4.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2049.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36',
        'Mozilla/5.0 (X11; OpenBSD i386) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1944.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2309.372 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2117.157 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1866.237 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.137 Safari/4E423F',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.116 Safari/537.36 Mozilla/5.0 (iPad; U; CPU OS 3_2 like Mac OS X; en-us) AppleWebKit/531.21.10 (KHTML, like Gecko) Version/4.0.4 Mobile/7B334b Safari/531.21.10',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.517 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1667.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1664.3 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1664.3 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.16 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1623.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.17 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/29.0.1547.62 Safari/537.36',
        'Mozilla/5.0 (X11; CrOS i686 4319.74.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/29.0.1547.57 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/29.0.1547.2 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1468.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1467.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1464.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1500.55 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.90 Safari/537.36',
        'Mozilla/5.0 (X11; NetBSD) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36',
        'Mozilla/5.0 (X11; CrOS i686 3912.101.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.116 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.60 Safari/537.17',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_2) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1309.0 Safari/537.17',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.15 (KHTML, like Gecko) Chrome/24.0.1295.0 Safari/537.15',
        'Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.14 (KHTML, like Gecko) Chrome/24.0.1292.0 Safari/537.14',
        'Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16',
        'Opera/9.80 (Windows NT 6.0) Presto/2.12.388 Version/12.14',
        'Mozilla/5.0 (Windows NT 6.0; rv:2.0) Gecko/20100101 Firefox/4.0 Opera 12.14',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0) Opera 12.14',
        'Opera/12.80 (Windows NT 5.1; U; en) Presto/2.10.289 Version/12.02',
        'Opera/9.80 (Windows NT 6.1; U; es-ES) Presto/2.9.181 Version/12.00',
        'Opera/9.80 (Windows NT 5.1; U; zh-sg) Presto/2.9.181 Version/12.00',
        'Opera/12.0(Windows NT 5.2;U;en)Presto/22.9.168 Version/12.00',
        'Opera/12.0(Windows NT 5.1;U;en)Presto/22.9.168 Version/12.00',
        'Mozilla/5.0 (Windows NT 5.1) Gecko/20100101 Firefox/14.0 Opera/12.0',
        'Opera/9.80 (Windows NT 6.1; WOW64; U; pt) Presto/2.10.229 Version/11.62',
        'Opera/9.80 (Windows NT 6.0; U; pl) Presto/2.10.229 Version/11.62',
        'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52',
        'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; de) Presto/2.9.168 Version/11.52',
        'Opera/9.80 (Windows NT 5.1; U; en) Presto/2.9.168 Version/11.51',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; de) Opera 11.51',
        'Opera/9.80 (X11; Linux x86_64; U; fr) Presto/2.9.168 Version/11.50',
        'Opera/9.80 (X11; Linux i686; U; hu) Presto/2.9.168 Version/11.50',
        'Opera/9.80 (X11; Linux i686; U; ru) Presto/2.8.131 Version/11.11',
        'Opera/9.80 (X11; Linux i686; U; es-ES) Presto/2.8.131 Version/11.11',
        'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/5.0 Opera 11.11',
        'Opera/9.80 (X11; Linux x86_64; U; bg) Presto/2.8.131 Version/11.10',
        'Opera/9.80 (Windows NT 6.0; U; en) Presto/2.8.99 Version/11.10',
        'Opera/9.80 (Windows NT 5.1; U; zh-tw) Presto/2.8.131 Version/11.10',
        'Opera/9.80 (Windows NT 6.1; Opera Tablet/15165; U; en) Presto/2.8.149 Version/11.1',
        'Opera/9.80 (X11; Linux x86_64; U; Ubuntu/10.10 (maverick); pl) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (X11; Linux i686; U; ja) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (X11; Linux i686; U; fr) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.1; U; zh-tw) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.1; U; sv) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.1; U; en-US) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.1; U; cs) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 6.0; U; pl) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 5.2; U; ru) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 5.1; U;) Presto/2.7.62 Version/11.01',
        'Opera/9.80 (Windows NT 5.1; U; cs) Presto/2.7.62 Version/11.01',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.2.13) Gecko/20101213 Opera/9.80 (Windows NT 6.1; U; zh-tw) Presto/2.7.62 Version/11.01',
        'Mozilla/5.0 (Windows NT 6.1; U; nl; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6 Opera 11.01',
        'Mozilla/5.0 (Windows NT 6.1; U; de; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6 Opera 11.01',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; de) Opera 11.01',
        'Opera/9.80 (X11; Linux x86_64; U; pl) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (X11; Linux i686; U; it) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.1; U; zh-cn) Presto/2.6.37 Version/11.00',
        'Opera/9.80 (Windows NT 6.1; U; pl) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.1; U; ko) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.1; U; fi) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.1; U; en-GB) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.1 x64; U; en) Presto/2.7.62 Version/11.00',
        'Opera/9.80 (Windows NT 6.0; U; en) Presto/2.7.39 Version/11.00',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1',
        'Mozilla/5.0 (Windows NT 6.3; rv:36.0) Gecko/20100101 Firefox/36.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10; rv:33.0) Gecko/20100101 Firefox/33.0',
        'Mozilla/5.0 (X11; Linux i586; rv:31.0) Gecko/20100101 Firefox/31.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:31.0) Gecko/20130401 Firefox/31.0',
        'Mozilla/5.0 (Windows NT 5.1; rv:31.0) Gecko/20100101 Firefox/31.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:29.0) Gecko/20120101 Firefox/29.0',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:25.0) Gecko/20100101 Firefox/29.0',
        'Mozilla/5.0 (X11; OpenBSD amd64; rv:28.0) Gecko/20100101 Firefox/28.0',
        'Mozilla/5.0 (X11; Linux x86_64; rv:28.0) Gecko/20100101  Firefox/28.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:27.3) Gecko/20130101 Firefox/27.3',
        'Mozilla/5.0 (Windows NT 6.2; Win64; x64; rv:27.0) Gecko/20121011 Firefox/27.0',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:25.0) Gecko/20100101 Firefox/25.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:25.0) Gecko/20100101 Firefox/25.0',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:24.0) Gecko/20100101 Firefox/24.0',
        'Mozilla/5.0 (Windows NT 6.0; WOW64; rv:24.0) Gecko/20100101 Firefox/24.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.8; rv:24.0) Gecko/20100101 Firefox/24.0',
        'Mozilla/5.0 (Windows NT 6.2; rv:22.0) Gecko/20130405 Firefox/23.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20130406 Firefox/23.0',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:23.0) Gecko/20131011 Firefox/23.0',
        'Mozilla/5.0 (Windows NT 6.2; rv:22.0) Gecko/20130405 Firefox/22.0',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:22.0) Gecko/20130328 Firefox/22.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:22.0) Gecko/20130405 Firefox/22.0',
        'Mozilla/5.0 (Microsoft Windows NT 6.2.9200.0); rv:22.0) Gecko/20130405 Firefox/22.0',
        'Mozilla/5.0 (Windows NT 6.2; Win64; x64; rv:16.0.1) Gecko/20121011 Firefox/21.0.1',
        'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:16.0.1) Gecko/20121011 Firefox/21.0.1',
        'Mozilla/5.0 (Windows NT 6.2; Win64; x64; rv:21.0.0) Gecko/20121011 Firefox/21.0.0',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:21.0) Gecko/20130331 Firefox/21.0',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (X11; Linux i686; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.2; WOW64; rv:21.0) Gecko/20130514 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.2; rv:21.0) Gecko/20130326 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20130401 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20130331 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20130330 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:21.0) Gecko/20130401 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:21.0) Gecko/20130328 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 5.1; rv:21.0) Gecko/20130401 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 5.1; rv:21.0) Gecko/20130331 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 5.1; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 5.0; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.8; rv:21.0) Gecko/20100101 Firefox/21.0',
        'Mozilla/5.0 (Windows NT 6.2; Win64; x64;) Gecko/20100101 Firefox/20.0',
        'Mozilla/5.0 (Windows x86; rv:19.0) Gecko/20100101 Firefox/19.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:6.0) Gecko/20100101 Firefox/19.0',
        'Mozilla/5.0 (Windows NT 6.1; rv:14.0) Gecko/20100101 Firefox/18.0.1',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:18.0)  Gecko/20100101 Firefox/18.0',
        'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:17.0) Gecko/20100101 Firefox/17.0.6',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; AS; rv:11.0) like Gecko',
        'Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0;  rv:11.0) like Gecko',
        'Mozilla/5.0 (compatible; MSIE 10.6; Windows NT 6.1; Trident/5.0; InfoPath.2; SLCC1; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729; .NET CLR 2.0.50727) 3gpp-gba UNTRUSTED/1.0',
        'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 7.0; InfoPath.3; .NET CLR 3.1.40767; Trident/6.0; en-IN)',
        'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0)',
        'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0)',
        'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/5.0)',
        'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/4.0; InfoPath.2; SV1; .NET CLR 2.0.50727; WOW64)',
        'Mozilla/5.0 (compatible; MSIE 10.0; Macintosh; Intel Mac OS X 10_7_3; Trident/6.0)',
        'Mozilla/4.0 (Compatible; MSIE 8.0; Windows NT 5.2; Trident/6.0)',
        'Mozilla/4.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/5.0)',
        'Mozilla/1.22 (compatible; MSIE 10.0; Windows 3.1)',
        'Mozilla/5.0 (Windows; U; MSIE 9.0; WIndows NT 9.0; en-US))',
        'Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 7.1; Trident/5.0)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; Media Center PC 6.0; InfoPath.3; MS-RTC LM 8; Zune 4.7)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; Media Center PC 6.0; InfoPath.3; MS-RTC LM 8; Zune 4.7',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 4.0; InfoPath.3; MS-RTC LM 8; .NET4.0C; .NET4.0E)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; chromeframe/12.0.742.112)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 2.0.50727; SLCC2; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 4.0; Tablet PC 2.0; InfoPath.3; .NET4.0C; .NET4.0E)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; yie8)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET CLR 1.1.4322; .NET4.0C; Tablet PC 2.0)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; FunWebProducts)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; chromeframe/13.0.782.215)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0; chromeframe/11.0.696.57)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0) chromeframe/10.0.648.205',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/4.0; GTB7.4; InfoPath.1; SV1; .NET CLR 2.8.52393; WOW64; en-US)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/5.0; chromeframe/11.0.696.57)',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.0; Trident/4.0; GTB7.4; InfoPath.3; SV1; .NET CLR 3.1.76908; WOW64; en-US)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0; GTB7.4; InfoPath.2; SV1; .NET CLR 3.3.69573; WOW64; en-US)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; InfoPath.1; SV1; .NET CLR 3.8.36217; WOW64; en-US)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; .NET CLR 2.7.58687; SLCC2; Media Center PC 5.0; Zune 3.4; Tablet PC 3.6; InfoPath.3)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.2; Trident/4.0; Media Center PC 4.0; SLCC1; .NET CLR 3.0.04320)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; SLCC1; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729; .NET CLR 1.1.4322)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; InfoPath.2; SLCC1; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729; .NET CLR 2.0.50727)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.1; SLCC1; .NET CLR 1.1.4322)',
        'Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 5.0; Trident/4.0; InfoPath.1; SV1; .NET CLR 3.0.4506.2152; .NET CLR 3.5.30729; .NET CLR 3.0.04506.30)',
        'Mozilla/5.0 (compatible; MSIE 7.0; Windows NT 5.0; Trident/4.0; FBSMTWB; .NET CLR 2.0.34861; .NET CLR 3.0.3746.3218; .NET CLR 3.5.33652; msn OptimizedIE8;ENUS)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.2; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; Media Center PC 6.0; InfoPath.2; MS-RTC LM 8)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; Media Center PC 6.0; InfoPath.2; MS-RTC LM 8',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; Media Center PC 6.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET4.0C)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; InfoPath.3; .NET4.0C; .NET4.0E; .NET CLR 3.5.30729; .NET CLR 3.0.30729; MS-RTC LM 8)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; InfoPath.2)',
        'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 3.0)',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A',
        'Mozilla/5.0 (iPad; CPU OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5355d Safari/8536.25',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_6_8) AppleWebKit/537.13+ (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/534.55.3 (KHTML, like Gecko) Version/5.1.3 Safari/534.53.10',
        'Mozilla/5.0 (iPad; CPU OS 5_1 like Mac OS X) AppleWebKit/534.46 (KHTML, like Gecko ) Version/5.1 Mobile/9B176 Safari/7534.48.3',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; de-at) AppleWebKit/533.21.1 (KHTML, like Gecko) Version/5.0.5 Safari/533.21.1',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_7; da-dk) AppleWebKit/533.21.1 (KHTML, like Gecko) Version/5.0.5 Safari/533.21.1',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; tr-TR) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; ko-KR) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; fr-FR) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; cs-CZ) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; ja-JP) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-US) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10_5_8; zh-cn) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10_5_8; ja-jp) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_7; ja-jp) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; zh-cn) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; sv-se) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; ko-kr) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; ja-jp) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; it-it) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; fr-fr) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; es-es) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; en-us) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; en-gb) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; de-de) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; sv-SE) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; ja-JP) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; de-DE) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; hu-HU) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-US) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; de-DE) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; ru-RU) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; ja-JP) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; it-IT) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_7; en-us) AppleWebKit/534.16+ (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; fr-ch) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_5; de-de) AppleWebKit/534.15+ (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_5; ar) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Android 2.2; Windows; U; Windows NT 6.1; en-US) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.3 Safari/533.19.4',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; zh-HK) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; tr-TR) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; nb-NO) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 6.0; fr-FR) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-TW) AppleWebKit/533.19.4 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; ru-RU) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_5_8; zh-cn) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5']
    headers = {'User-Agent': random.choice(user_agent_list)}
    return headers


class ConnectSqlite:

    def __init__(self, dbName="./sqlite3Test.db"):
        """
        初始化连接--使用完记得关闭连接
        :param dbName: 连接库名字，注意，以'.db'结尾
        """
        self._conn = sqlite3.connect(dbName, timeout=3, isolation_level=None, check_same_thread=False)
        self._cur = self._conn.cursor()
        self._time_now = "[" + sqlite3.datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S') + "]"

    def close_con(self):
        """
        关闭连接对象--主动调用
        :return:
        """
        self._cur.close()
        self._conn.close()

    def create_tabel(self, sql):
        """
        创建表初始化
        :param sql: 建表语句
        :return: True is ok
        """
        try:
            self._cur.execute(sql)
            self._conn.commit()
            return True
        except Exception as e:
            print(self._time_now, "[CREATE TABLE ERROR]", e)
            return False

    def drop_table(self, table_name):
        """
        删除表
        :param table_name: 表名
        :return:
        """
        try:
            self._cur.execute('DROP TABLE {0}'.format(table_name))
            self._conn.commit()
            return True
        except Exception as e:
            print(self._time_now, "[DROP TABLE ERROR]", e)
            return False

    def delete_table(self, sql):
        """
        删除表记录
        :param sql:
        :return: True or False
        """
        try:
            if 'DELETE' in sql.upper():
                self._cur.execute(sql)
                self._conn.commit()
                return True
            else:
                print(self._time_now, "[EXECUTE SQL IS NOT DELETE]")
                return False
        except Exception as e:
            print(self._time_now, "[DELETE TABLE ERROR]", e)
            return False

    def fetchall_table(self, sql, limit_flag=True):
        """
        查询所有数据
        :param sql:
        :param limit_flag: 查询条数选择，False 查询一条，True 全部查询
        :return:
        """
        try:
            self._cur.execute(sql)
            war_msg = self._time_now + ' The [{}] is empty or equal None!'.format(sql)
            if limit_flag is True:
                r = self._cur.fetchall()
                return r if len(r) > 0 else war_msg
            elif limit_flag is False:
                r = self._cur.fetchone()
                return r if len(r) > 0 else war_msg
        except Exception as e:
            print(self._time_now, "[SELECT TABLE ERROR]", e)

    def insert_update_table(self, sql):
        """
        插入/更新表记录
        :param sql:
        :return:
        """
        try:
            self._cur.execute(sql)
            self._conn.commit()
            return True
        except Exception as e:
            print(self._time_now, "[INSERT/UPDATE TABLE ERROR]", e)
            return False

    def insert_table_many(self, sql, value):
        """
        插入多条记录
        :param sql:
        :param value: list:[(),()]
        :return:
        """
        try:
            self._cur.executemany(sql, value)
            self._conn.commit()
            return True
        except Exception as e:
            print(self._time_now, "[INSERT MANY TABLE ERROR]", e)
            return False


out_path = yaml_data.get('OUT_FILE_PATH')
out_path = out_path if out_path else 'data/网上最新价格.xlsx'
try:
    wb = load_workbook(out_path)
except FileNotFoundError as e:
    wb = Workbook()
conn = ConnectSqlite("./sqlite3Ip.db")
notes_row = 2


@atexit.register
def exit_handle():
    print('匹配到第 {} 件商品结束'.format(notes_row))
    conn.insert_update_table('''UPDATE notes SET number={0} WHERE id={1}'''.format(notes_row, '520'))
    wb.save(out_path)
    conn.close_con()


def on_close(sig):
    conn.insert_update_table('''UPDATE notes SET number={0} WHERE id={1}'''.format(notes_row, '520'))
    wb.save(out_path)
    conn.close_con()
    sys.exit()


win32api.SetConsoleCtrlHandler(on_close, True)


class Main:

    def __init__(self):
        pass

    def getIp(self):
        ip_list = conn.fetchall_table('SELECT * FROM proxyip;')
        if isinstance(ip_list, list) and len(ip_list) <= 1:
            self.getip()
            while True:
                ip_list = conn.fetchall_table('SELECT * FROM proxyip;')
                if len(ip_list) >= 10:
                    break
                time.sleep(10)
        elif isinstance(ip_list, str):
            self.getip()
            while True:
                ip_list = conn.fetchall_table('SELECT * FROM proxyip;')
                if len(ip_list) >= 10:
                    break
                time.sleep(10)
        if isinstance(ip_list, list) and len(ip_list) > 0:
            return random.choice(ip_list)
        else:
            raise RuntimeError('程序异常结束')

    def requests_process(self, row, sku, commodity_name):
        headers = getheaders()  # 定制请求头
        ip_tuple = self.getIp()
        if not isinstance(ip_tuple, tuple):
            return ['Ip代理获取失败', 0]
        proxies = {"http": "http://" + ip_tuple[0], "https": "http://" + ip_tuple[0]}  # 代理ip
        requests.adapters.DEFAULT_RETRIES = 10
        s = requests.session()
        s.keep_alive = False
        try:
            response = requests.get(url=search_url.format(sku), proxies=proxies, headers=headers, timeout=10, verify=False)
            if int(response.status_code) == 200:
                soup = BeautifulSoup(response.text, 'lxml')
                all_span = soup.select('#searchTabPrdList .imgType .listUl .productMd .price span')
                if len(all_span) > 1:
                    return ['商品搜索条数错误', 0]
                elif len(all_span) == 1:
                    match = pattern.findall(all_span[0].get_text())
                    if match:
                        return ['搜索成功', re.search(r'\d+(\.\d+)?', all_span[0].get_text()).group()]
                    else:
                        all_strong = soup.select('#searchTabPrdList .imgType .listUl .productMd .discount strong')
                        return ['搜索成功', re.search(r'\d+(\.\d+)?', all_strong[0].get_text()).group()]
                else:
                    return ['商品没有搜到', 0]
            elif int(response.status_code) == 403:
                print('第 {0} 件商品搜索失败---sku号为：{1}--商品名称为：{2}--错误为：403错误'.format(row, sku, commodity_name))
                conn.delete_table('''DELETE FROM proxyip WHERE ip_port='{0}';'''.format(ip_tuple[0]))
                return self.requests_process(row, sku, commodity_name)
            else:
                return ['商品搜索失败', 0]
        except Exception:
            print('第 {0} 件商品匹配失败---sku号为：{1}--商品名称为：{2}--错误为：超时/代理错误'.format(row, sku, commodity_name))
            conn.delete_table('''DELETE FROM proxyip WHERE ip_port='{0}';'''.format(ip_tuple[0]))
            return self.requests_process(row, sku, commodity_name)

    def process(self, rs, ws):
        global notes_row
        start_row = 2
        start_row_list = conn.fetchall_table('''select number from notes where id = '520';''')
        if len(start_row_list) > 0 and start_row_list[0][0]:
            start_row = start_row_list[0][0]
            notes_row = start_row
        if start_row == 2:
            for n in range(1, ws.max_row + 1):
                ws.delete_rows(n)
        wb.save(out_path)
        if ws.max_row <= 1:
            data_list = ['序号', 'sku', '品牌', '名称', '原价', '搜索结果', '网上价格']
            ws.append(data_list)
            wb.save(out_path)
        for row in range(start_row, rs.max_row + 1):
            sku_column = yaml_data.get('SKU_COLUMN')  # sku
            sku_column = sku_column if sku_column else 1
            brand_column = yaml_data.get('BRAND_COLUMN')  # 品牌
            brand_column = brand_column if brand_column else 2
            commodity_name_column = yaml_data.get('COMMODITY_NAME_COLUMN')  # 商品名称
            commodity_name_column = commodity_name_column if commodity_name_column else 3
            original_price_column = yaml_data.get('ORIGINAL_PRICE_COLUMN')  # 原价
            original_price_column = original_price_column if original_price_column else 4
            sku = rs.cell(row=row, column=sku_column).value
            brand = rs.cell(row=row, column=brand_column).value
            commodity_name = rs.cell(row=row, column=commodity_name_column).value
            original_price = rs.cell(row=row, column=original_price_column).value
            if sku and original_price:
                data = self.requests_process(row, sku, commodity_name)
                print('第 {0} 件商品处理结果为：{1}---sku号为：{2}---商品名称为：{3}---表格价格：{4}---网上价格为：{5}'.format(
                    row, data[0], sku, commodity_name, original_price, data[1]))
                data_list = [row, sku, brand, commodity_name, original_price, data[0], data[1]]
                ws.append(data_list)
                wb.save(out_path)
            else:
                print('第 {0} 件商品匹配失败---sku号为：{1}---商品名称为：{2}  --->sku非法或者价格非法'.format(row, sku, commodity_name))
                data_list = [row, sku, brand, commodity_name, original_price, 'sku非法或者价格非法', 0]
                ws.append(data_list)
                wb.save(out_path)
            notes_row = row
        print('总共匹配了 {0} 件商品价格'.format(notes_row))

    # -----------------------------------------------------检查ip是否可用----------------------------------------------------
    def checkip(self, ip):
        headers = getheaders()  # 定制请求头
        proxies = {"http": "http://" + ip, "https": "http://" + ip}  # 代理ip
        requests.adapters.DEFAULT_RETRIES = 3
        thisIP = "".join(ip.split(":")[0:1])
        try:
            response = requests.get(url=targeturl, proxies=proxies, headers=headers, timeout=5)
            if thisIP in response.text:
                return True
            else:
                return False
        except Exception:
            return False

    # -------------------------------------------------------获取代理方法----------------------------------------------------
    # 免费代理 XiciDaili
    def findip(self, type, pagenum):  # ip类型,页码,目标url,存放ip的路径
        list = {'1': 'http://www.xicidaili.com/wn/',  # xicidaili国内https代理
                '2': 'http://www.xicidaili.com/nn/',  # xicidaili国内高匿代理
                '3': 'http://www.xicidaili.com/nt/',  # xicidaili国内普通代理
                '4': 'http://www.xicidaili.com/wt/'}  # xicidaili国外http代理
        url = list[str(type)] + str(pagenum)  # 配置url
        headers = getheaders()  # 定制请求头
        try:
            html = requests.get(url=url, headers=headers, timeout=5).text
            soup = BeautifulSoup(html, 'lxml')
            all = soup.find_all('tr', class_='odd')
            for i in all:
                t = i.find_all('td')
                ip = t[1].text + ':' + t[2].text
                is_avail = self.checkip(ip)
                if is_avail:
                    sql = """INSERT INTO proxyip VALUES ('{0}');""".format(ip)
                    print('代理Ip: {0} 插入成功'.format(ip) if conn.insert_update_table(sql) else '代理Ip: {0} 插入失败'.format(ip))
        except Exception:
            print('代理Ip请求失败，可能Ip被禁止访问，请刷新网络Ip重启exe文件')

    # -----------------------------------------------------多线程抓取ip入口---------------------------------------------------
    def getip(self):
        threads = []
        for type in range(4):  # 四种类型ip,每种类型取前三页,共12条线程
            for pagenum in range(5):
                t = threading.Thread(target=self.findip, args=(type + 1, pagenum + 1))
                threads.append(t)
        for s in threads:  # 开启多线程爬取
            s.start()

    # -------------------------------------------------------读取文件execl-----------------------------------------------------------
    def readFile(self):
        ws = wb.active
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['A'].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions['C'].width = 36
        ws.column_dimensions['C'].alignment = Alignment(horizontal='center', vertical='center')
        file_path = yaml_data.get('FILE_PATH')
        file_path = file_path if file_path else 'data/欧美韩免原价.xlsx'
        rb = load_workbook(file_path)
        sheets = rb.sheetnames
        sheet = sheets[0]
        rs = rb[sheet]
        self.process(rs, ws)


# -------------------------------------------------------启动-----------------------------------------------------------
if __name__ == '__main__':
    sql = '''CREATE TABLE `proxyip` (
                      `ip_port` VARCHAR(25) DEFAULT NULL PRIMARY KEY
                    )'''
    print('创建代理表成功' if conn.create_tabel(sql) else '创建代理表失败')
    sql1 = '''CREATE TABLE `notes` (
                      `id` VARCHAR(5) DEFAULT NULL PRIMARY KEY,
                      `number` int(6) DEFAULT NULL
                    )'''
    if conn.create_tabel(sql1):
        print('创建记录表成功')
        conn.insert_update_table('''INSERT INTO notes VALUES ('520', 2);''')
    else:
        print('创建记录表失败')
    ip_list = conn.fetchall_table('SELECT * FROM proxyip;')
    m = Main()
    if isinstance(ip_list, list) and len(ip_list) <= 10:
        m.getip()
    elif isinstance(ip_list, str):
        m.getip()
    while True:
        ip_list = conn.fetchall_table('SELECT * FROM proxyip;')
        if isinstance(ip_list, list) and len(ip_list) >= 5:
            break
        time.sleep(3)
        if isinstance(ip_list, list):
            print('---Ip代理数量不够，正常5个，等待数量满足开始匹配，当前代理Ip个数为：（{0}）---'.format(len(ip_list)))
        else:
            print('---Ip代理数量不够，正常5个，等待数量满足开始匹配，当前代理Ip个数为：（{0}）---'.format(0))
    m.readFile()
    input('点击右上角关闭')
    while True:
        time.sleep(60)
