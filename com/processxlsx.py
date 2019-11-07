# !/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import sqlite3
import atexit
import yaml
import time
import sys
import os

notes_row = 2
yamlPath = 'config.yaml'
_yaml = open(yamlPath, 'r', encoding='utf-8')
cont = _yaml.read()
yaml_data = yaml.load(cont, Loader=yaml.FullLoader)
sys.path.append(os.path.abspath(os.path.dirname(__file__) + '/' + '..'))
sys.path.append("..")
from com.ConnectSqlite import ConnectSqlite

conn = ConnectSqlite("./.SqliteData.db")
@atexit.register
def exit_handle():
    conn.insert_update_table('''UPDATE notes SET number={0} WHERE id={1}'''.format(notes_row, '520'))
    conn.close_con()
    print('数据转存数据库结束')


class Xlsx:

    def __init__(self, file_path, start_row):
        self.file_path = file_path
        self.start_row = start_row
        start_row_list = conn.fetchall_table('''select number from notes where id = '520';''')
        if len(start_row_list) > 0 and start_row_list[0][0]:
            self.start_row = start_row_list[0][0]
        self.dic = {}

    def getdata(self, row, rs):
        sku_column = yaml_data.get('SKU_COLUMN')  # sku
        sku_column = sku_column if sku_column else 1
        brand_column = yaml_data.get('BRAND_COLUMN')  # 品牌
        brand_column = brand_column if brand_column else 2
        commodity_name_column = yaml_data.get('COMMODITY_NAME_COLUMN')  # 商品名称
        commodity_name_column = commodity_name_column if commodity_name_column else 3
        original_price_column = yaml_data.get('ORIGINAL_PRICE_COLUMN')  # 原价
        original_price_column = original_price_column if original_price_column else 4
        sku = rs.cell(row=row, column=sku_column).value
        brand = rs.cell(row=row, column=brand_column).value
        commodity_name = rs.cell(row=row, column=commodity_name_column).value
        original_price = rs.cell(row=row, column=original_price_column).value
        return [sku, brand, commodity_name, original_price]

    def wirtesqlite(self, rs):
        global notes_row
        max_row = rs.max_row+1
        for row in range(self.start_row, max_row):
            data = self.getdata(row, rs)
            if data[0]:
                sql = """INSERT INTO originaldata VALUES ({0}, {1}, {2});""".format(data[0], data[3], 0)
                if conn.insert_update_table(sql):
                    print('第 {0} 条插入成功'.format(row))
                else:
                    print('第 {0} 条插入失败'.format(row))
            else:
                print('第 {0} 条插入失败'.format(row))
            notes_row = row
        # sql = """INSERT INTO originaldata VALUES (?, ?, ?)"""
        # row_list = [n for n in range(1, rs.max_row + 1)]
        # row_list = [row_list[i:i+100]
        #              for i in range(0, len(row_list), 100)]
        # for row in row_list:
        #     value = []
        #     for r in row:
        #         data = self.getdata(r, rs)
        #         value.append((data[0], data[3], 0))
        #         # rs.delete_rows(r)
        #     print(conn.insert_table_many(sql, value))
        
    def readfile(self):
        rb = load_workbook(self.file_path)
        sheets = rb.sheetnames
        sheet = sheets[0]
        rs = rb[sheet]
        self.wirtesqlite(rs)
        rb.save(self.file_path)


if __name__ == "__main__":
    start = time.time()
    sql = '''CREATE TABLE `originaldata` (
            `sku` VARCHAR(12) DEFAULT NULL PRIMARY KEY,
            `original_price` VARCHAR(9) DEFAULT NULL,
            `code` int(1) DEFAULT NULL
            )'''
    print('创建原始数据表成功' if conn.create_tabel(sql) else '创建原始数据表失败')
    sql = '''CREATE TABLE `notes` (
                      `id` VARCHAR(5) DEFAULT NULL PRIMARY KEY,
                      `number` int(6) DEFAULT NULL
                    )'''
    if conn.create_tabel(sql):
        print('创建记录表成功')
        conn.insert_update_table('''INSERT INTO notes VALUES ('520', 2);''')
    else:
        print('创建记录表失败')
    file_path = yaml_data.get('FILE_PATH')
    file_path = file_path if file_path else 'data/欧美韩免原价.xlsx'
    start_row = yaml_data.get('START_ROW')
    start_row = start_row if start_row else 2
    x = Xlsx(file_path, start_row)
    x.readfile()
    print("运行完毕，总用时:{}".format(time.time() - start))
